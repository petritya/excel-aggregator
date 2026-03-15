from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
import pandas as pd
import io
import zipfile
from typing import List, Tuple
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

app = FastAPI(title="Excel összesítő")


@app.get("/", response_class=HTMLResponse)
def home():
    return """
<!doctype html>
<html lang="hu">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Excel összesítő</title>
  <style>
    :root { color-scheme: dark; }
    body {
      margin: 0;
      font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial;
      background: #0b1020;
      color: #e6e8ee;
    }
    .wrap { max-width: 820px; margin: 0 auto; padding: 28px 18px; }
    .card {
      background: rgba(255,255,255,0.06);
      border: 1px solid rgba(255,255,255,0.12);
      border-radius: 16px;
      padding: 18px;
      box-shadow: 0 12px 40px rgba(0,0,0,0.35);
    }
    h1 { font-size: 20px; margin: 0 0 10px; }
    p { margin: 8px 0; line-height: 1.5; color: rgba(230,232,238,0.9); }
    .row { display: flex; gap: 12px; flex-wrap: wrap; align-items: center; margin-top: 14px; }
    .file {
      flex: 1;
      min-width: 240px;
      padding: 10px 12px;
      background: rgba(0,0,0,0.25);
      border: 1px dashed rgba(255,255,255,0.18);
      border-radius: 12px;
    }
    input[type="file"] { width: 100%; }
    button {
      border: 0;
      border-radius: 12px;
      padding: 10px 14px;
      background: #6d5efc;
      color: white;
      font-weight: 700;
      cursor: pointer;
    }
    button:disabled { opacity: 0.6; cursor: not-allowed; }
    .status { margin-top: 12px; padding: 10px 12px; border-radius: 12px; display:none; }
    .status.ok { display:block; background: rgba(45, 212, 191, 0.15); border: 1px solid rgba(45, 212, 191, 0.35); }
    .status.err { display:block; background: rgba(248, 113, 113, 0.12); border: 1px solid rgba(248, 113, 113, 0.35); }
    .spinner {
      width: 14px; height: 14px;
      border: 2px solid rgba(255,255,255,0.25);
      border-top-color: rgba(255,255,255,0.95);
      border-radius: 50%;
      display:inline-block;
      animation: spin 0.8s linear infinite;
      vertical-align: -2px;
      margin-right: 8px;
    }
    @keyframes spin { to { transform: rotate(360deg); } }
    .footer { margin-top: 10px; opacity: 0.7; font-size: 12px; }
    a { color: #bcb6ff; }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <h1>Excel összesítő</h1>
      <p>
        Tölts fel több <b>.xlsx</b> fájlt vagy egy <b>.zip</b> csomagot.
        A rendszer automatikusan kicsomagolja a ZIP-et, eldobja az első 5 oszlopot,
        majd cikkszám + név alapján összeadja a darabszámot.
      </p>

      <div class="row">
        <div class="file">
          <input id="files" type="file" multiple accept=".xlsx,.zip" />
        </div>
        <button id="run" disabled>Összesítés indítása</button>
      </div>

      <div id="status" class="status"></div>

      <div class="footer">
        API teszt: <a href="/docs">/docs</a>
      </div>
    </div>
  </div>

  <script>
    const filesEl = document.getElementById("files");
    const runBtn = document.getElementById("run");
    const statusEl = document.getElementById("status");

    function setStatusOk(message) {
      statusEl.className = "status ok";
      statusEl.innerHTML = message;
    }

    function setStatusErr(message) {
      statusEl.className = "status err";
      statusEl.textContent = message;
    }

    filesEl.addEventListener("change", () => {
      runBtn.disabled = !(filesEl.files && filesEl.files.length);

      if (filesEl.files.length) {
        setStatusOk("Kiválasztva: " + filesEl.files.length + " fájl");
      } else {
        statusEl.className = "status";
        statusEl.innerHTML = "";
      }
    });

    runBtn.addEventListener("click", async () => {
      runBtn.disabled = true;
      setStatusOk('<span class="spinner"></span>Fájlok feldolgozása és összesítés folyamatban...');

      const fd = new FormData();
      for (const f of filesEl.files) {
        fd.append("files", f);
      }

      try {
        const res = await fetch("/tools/excel/aggregate", {
          method: "POST",
          body: fd
        });

        if (!res.ok) {
          const msg = await res.text();
          setStatusErr("Hiba: " + msg);
          runBtn.disabled = false;
          return;
        }

        const zipExcelCount = res.headers.get("X-Zip-Excel-Count") || "0";
        const totalExcelCount = res.headers.get("X-Total-Excel-Count") || "0";
        const originalRows = res.headers.get("X-Original-Row-Count") || "0";
        const aggregatedRows = res.headers.get("X-Aggregated-Row-Count") || "0";

        const blob = await res.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "osszesitett_struktura_tisztitott.xlsx";
        document.body.appendChild(a);
        a.click();
        a.remove();
        URL.revokeObjectURL(url);

        filesEl.value = "";
        runBtn.disabled = true;

        setStatusOk(
          "Kész! Az összesített fájl letöltődött.<br>" +
          "Feldolgozott Excel fájlok száma: <b>" + totalExcelCount + "</b><br>" +
          "ZIP-ből kinyert Excel fájlok száma: <b>" + zipExcelCount + "</b><br>" +
          "Eredeti sorok száma: <b>" + originalRows + "</b><br>" +
          "Összesített sorok száma: <b>" + aggregatedRows + "</b>"
        );
      } catch (e) {
        setStatusErr("Hálózati hiba történt.");
        runBtn.disabled = false;
      }
    });
  </script>
</body>
</html>
    """


def read_uploaded_excels(files: List[UploadFile]) -> Tuple[List[pd.DataFrame], int, int]:
    df_list: List[pd.DataFrame] = []
    zip_excel_count = 0
    total_excel_count = 0

    for uploaded_file in files:
        filename = (uploaded_file.filename or "").lower()
        content = uploaded_file.file.read()

        if not content:
            continue

        if filename.endswith(".xlsx"):
            try:
                df = pd.read_excel(io.BytesIO(content))
                df_list.append(df)
                total_excel_count += 1
            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Hibás Excel fájl: {uploaded_file.filename} ({str(e)})"
                )

        elif filename.endswith(".zip"):
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as zip_file:
                    for name in zip_file.namelist():
                        lower_name = name.lower()

                        if name.endswith("/") or "__macosx" in lower_name:
                            continue

                        if lower_name.endswith(".xlsx"):
                            with zip_file.open(name) as extracted_file:
                                file_bytes = extracted_file.read()
                                try:
                                    df = pd.read_excel(io.BytesIO(file_bytes))
                                    df_list.append(df)
                                    zip_excel_count += 1
                                    total_excel_count += 1
                                except Exception as e:
                                    raise HTTPException(
                                        status_code=400,
                                        detail=f"Hibás Excel a ZIP-ben: {name} ({str(e)})"
                                    )
            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Hibás ZIP fájl: {uploaded_file.filename} ({str(e)})"
                )

        else:
            raise HTTPException(
                status_code=400,
                detail=f"Nem támogatott fájltípus: {uploaded_file.filename}. Csak .xlsx és .zip engedélyezett."
            )

    if not df_list:
        raise HTTPException(
            status_code=400,
            detail="Nem található feldolgozható Excel fájl a feltöltésben."
        )

    return df_list, zip_excel_count, total_excel_count


def format_worksheet(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    alt_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if row_idx % 2 == 0:
                cell.fill = alt_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        adjusted_width = min(max(max_length + 2, 12), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    ws.row_dimensions[1].height = 22


def aggregate_dataframes(df_list: List[pd.DataFrame]) -> Tuple[bytes, int, int]:
    combined_df = pd.concat(df_list, ignore_index=True)
    original_row_count = len(combined_df)

    if combined_df.shape[1] <= 5:
        raise HTTPException(
            status_code=400,
            detail="Túl kevés oszlop: nem tudom eldobni az első 5 oszlopot."
        )

    combined_df = combined_df.iloc[:, 5:]
    original_columns = combined_df.columns.tolist()

    if len(original_columns) <= 6:
        raise HTTPException(
            status_code=400,
            detail="Nem találom a darabszám oszlopot a várt pozícióban."
        )

    cikkszam_col = combined_df.columns[0]
    nev_col = combined_df.columns[1]
    darab_col = combined_df.columns[6]

    agg_map = {col: "first" for col in combined_df.columns}
    agg_map[darab_col] = "sum"

    aggregated_df = (
        combined_df
        .groupby([cikkszam_col, nev_col], as_index=False)
        .agg(agg_map)
    )

    aggregated_df = aggregated_df[original_columns]
    aggregated_row_count = len(aggregated_df)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        aggregated_df.to_excel(writer, sheet_name="Osszesites", index=False)
        ws = writer.book["Osszesites"]
        format_worksheet(ws)

    output.seek(0)
    return output.getvalue(), original_row_count, aggregated_row_count


@app.post("/tools/excel/aggregate")
async def excel_aggregate(files: List[UploadFile] = File(...)):
    if not files:
        raise HTTPException(status_code=400, detail="Nincs feltöltött fájl.")

    df_list, zip_excel_count, total_excel_count = read_uploaded_excels(files)
    xlsx_bytes, original_row_count, aggregated_row_count = aggregate_dataframes(df_list)

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="osszesitett_struktura_tisztitott.xlsx"',
            "X-Zip-Excel-Count": str(zip_excel_count),
            "X-Total-Excel-Count": str(total_excel_count),
            "X-Original-Row-Count": str(original_row_count),
            "X-Aggregated-Row-Count": str(aggregated_row_count),
        },
    )